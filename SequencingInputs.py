from ast import Continue
from asyncio.windows_events import NULL
from distutils import command
from distutils.command.clean import clean
from distutils.dir_util import remove_tree
from locale import format_string
from msilib import type_key
from operator import countOf
from pickle import APPEND
from queue import Empty
from re import A
from ssl import Options
import string
import tkinter as tk
from tokenize import group
from traceback import print_list
from turtle import color
from xml.etree.ElementTree import tostringlist
import openpyxl
from tkinter import *
from tkinter import ttk
import os
import webbrowser
import math
import decimal
import time
from datetime import datetime, timedelta
from functools import partial

from PIL import ImageTk, Image


import checklistmain
import SequencingReport
import amzl_requests



root = tk.Tk()
apps = []

#set title to the window
root.title("PM Routing - Sequencing inputs")

#user login
user_login = os.getenv("username")
print(user_login)


root.iconbitmap(r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\PM_Dashboard\Sequencing inputs\centralops.ico')


#function when pressing RTS button
def excelfileRTS():
    
    top = Toplevel()
    top.title() 
    global minutesRTS #it must be a global variable to be displayed otherwise it wont work
    global minutes_listRTS
    global dep_time_listRTS
    global SIMRTSinfo
    

    global DS
    DS = DSentry.get()
    DSentry.delete(0, END) #delete DS input when clicking RTS button
    DSupper = DS.upper()
    top.title(DSupper)
    
    #title in the new window ( RTS + DS)
    frametitle = LabelFrame(top, padx=1, pady=1)
    frametitle.grid(row=4, columnspan=11, \
                 padx=1, pady=1, ipadx=1, ipady=1)
    titlelabel = Label(frametitle, text = "RTS " + DSupper, anchor= CENTER,  font = "Helvetica 15", padx=1, pady=1).grid(row=4, column=0, columnspan=1, pady=1, padx=1)

    #Setting frames for DSP minutes and Flex minutes
    
    #for sequencing minutes
    frame2 = LabelFrame(top, padx=5, pady=5, text="Minutes", font = "16")
    frame2.grid(row=5, columnspan=11, \
                 padx=5, pady=5, ipadx=5, ipady=5)
    
    #for SIMs
    frame0 = LabelFrame(top, padx=5, pady=5, text="SIMs", font = "16") 
    frame0.grid(row=13, columnspan=11, \
                 padx=5, pady=5, ipadx=5, ipady=5)
    
    
    
    #Reading Excel file
    
    # setting directory where file is located
    
    path = "//ant/dept-eu/TBA/UK/Business Analyses/CentralOPS/PM Shift/DHP1/PM_Dashboard/Sequencing inputs/Database.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True) #load workbook and telling Python to read only
    ws = wb.worksheets[0] #sheet number 1 is "RTS"
    wES = wb.worksheets[4] #sheet number 2 is "Adhoc" only for ES

    #looking for minutes depending on dep time            
    search_str = DS #we want to search for the DS specified
    minutes_listRTS = [] #the output of the minutes in list format
    dep_time_listRTS = []
    service_type_listRTS = []

    minutes_listRTS2 = [] #the output of the minutes in list format
    dep_time_listRTS2 = []
    service_type_listRTS2 = []

    minutes_listRTS2prescheduled = [] #the output of the minutes in list format
    dep_time_listRTS2prescheduled = []
    
    minutes_listRTS1prescheduled = [] #the output of the minutes in list format
    dep_time_listRTS1prescheduled = []
    

    range1 = ws._cells_by_col(1,1,5,10000) # Defaults to whole sheet (to search)
    range2 = wES.iter_rows()
    
    #for DS without flex
    for row in range2:
        for cell in row:
            if (cell.value == search_str.upper()):
                minutesRTS0 = ws.cell(row=3, column=7).value #get the value from column 7 and the row where DS has been found
                minutesRTS1 = ws.cell(row=4, column=7).value
                minutesRTS2 = ws.cell(row=5, column=7).value
                minutesRTS3 = ws.cell(row=6, column=7).value
                minutesRTS4 = ws.cell(row=7, column=7).value
                minutesRTS5 = ws.cell(row=8, column=7).value
                minutesRTS6 = ws.cell(row=8, column=7).value



                minutes_listRTS.append(minutesRTS0)
                minutes_listRTS.append(minutesRTS1)
                minutes_listRTS.append(minutesRTS2)
                minutes_listRTS.append(minutesRTS3)
                minutes_listRTS.append(minutesRTS4)
                minutes_listRTS.append(minutesRTS5)
                minutes_listRTS.append(minutesRTS6)

                #append the result to the list
                #print(minutes_listRTS)
                
                deptimesRTS0 = ws.cell(row=3, column=6).value
                deptimesRTS1 = ws.cell(row=4, column=6).value
                deptimesRTS2 = ws.cell(row=5, column=6).value
                deptimesRTS3 = ws.cell(row=6, column=6).value
                deptimesRTS4 = ws.cell(row=7, column=6).value
                deptimesRTS5 = ws.cell(row=8, column=6).value
                deptimesRTS6 = ws.cell(row=9, column=6).value

                dep_time_listRTS.append(deptimesRTS0)
                dep_time_listRTS.append(deptimesRTS1)
                dep_time_listRTS.append(deptimesRTS2)
                dep_time_listRTS.append(deptimesRTS3)
                dep_time_listRTS.append(deptimesRTS4)
                dep_time_listRTS.append(deptimesRTS5)
                dep_time_listRTS.append(deptimesRTS6)
                

                #print(dep_time_listRTS)

                servicetypeRTS0 = ws.cell(row=3, column=4).value
                servicetypeRTS1 = ws.cell(row=4, column=4).value
                servicetypeRTS2 = ws.cell(row=5, column=4).value
                servicetypeRTS3 = ws.cell(row=6, column=4).value
                servicetypeRTS4 = ws.cell(row=7, column=4).value
                servicetypeRTS5 = ws.cell(row=8, column=4).value
                servicetypeRTS6 = ws.cell(row=9, column=4).value
                
                service_type_listRTS.append(servicetypeRTS0)
                service_type_listRTS.append(servicetypeRTS1)
                service_type_listRTS.append(servicetypeRTS2)
                service_type_listRTS.append(servicetypeRTS3)
                service_type_listRTS.append(servicetypeRTS4)
                service_type_listRTS.append(servicetypeRTS5)
                service_type_listRTS.append(servicetypeRTS6)
            

            #for DS with flex enabled
            else:
                for row in range1:
                    for cell in row:
                        if (cell.value == search_str.upper()):
                            minutesRTS = ws.cell(row=cell.row, column=7).value #get the value from column 7 and the row where DS has been found
                            minutes_listRTS.append(minutesRTS) #append the result to the list
                            #RTS2
                            minutesRTS2 = ws.cell(row=cell.row, column=13).value #get the value from column 7 and the row where DS has been found
                            minutes_listRTS2.append(minutesRTS2)
                            #RTS2 prescheduled
                            minutesRTS2prescheduled = ws.cell(row=cell.row, column=23).value #get the value from column 7 and the row where DS has been found
                            minutes_listRTS2prescheduled.append(minutesRTS2prescheduled)
                            #RTS1 prescheduled
                            minutesRTS1prescheduled = ws.cell(row=cell.row, column=18).value #get the value from column 7 and the row where DS has been found
                            minutes_listRTS1prescheduled.append(minutesRTS1prescheduled)
                            



                            deptimesRTS = ws.cell(row=cell.row, column=6).value
                            dep_time_listRTS.append(deptimesRTS)
                            #RTS2
                            deptimesRTS2 = ws.cell(row=cell.row, column=12).value
                            dep_time_listRTS2.append(deptimesRTS2)
                            #RTS2 prescheduled
                            deptimesRTS2prescheduled = ws.cell(row=cell.row, column=22).value
                            dep_time_listRTS2prescheduled.append(deptimesRTS2prescheduled)
                            #RTS1 prescheduled
                            deptimesRTS1prescheduled = ws.cell(row=cell.row, column=17).value
                            dep_time_listRTS1prescheduled.append(deptimesRTS1prescheduled)


                            servicetypeRTS = ws.cell(row=cell.row, column=4).value
                            service_type_listRTS.append(servicetypeRTS)
                            #print(service_type_listRTS)
                            servicetypeRTS2 = ws.cell(row=cell.row, column=11).value
                            service_type_listRTS2.append(servicetypeRTS2)
                            
                


    #setting up the table to display the result
    set = ttk.Treeview(frame2)
    set.grid(row=7, column=0, columnspan=9, rowspan=2, pady=10, padx=10) #where the table will be located

    set['columns']= ('dispatch_time', 'minutes','service_type') #columns IDs
    set.column("#0", width=0,  stretch=NO)
    set.column("dispatch_time",anchor=CENTER, width=140)
    set.column("minutes",anchor=CENTER, width=80)
    set.column("service_type",anchor=CENTER, width=180)
    

    #name and formatting for headers
    set.heading("#0",text="",anchor=CENTER)
    set.heading("dispatch_time",text="Dispatch Time",anchor=CENTER)
    set.heading("minutes",text="Minutes",anchor=CENTER)
    set.heading("service_type",text="Service Type",anchor=CENTER)
    

    #data in the table
    data  = [
        [[dep_time_listRTS[0]], [minutes_listRTS[0]],[service_type_listRTS[0]]],
        [[dep_time_listRTS[1]], [minutes_listRTS[1]],[service_type_listRTS[1]]],
        [[dep_time_listRTS[2]], [minutes_listRTS[2]],[service_type_listRTS[2]]],
        [[dep_time_listRTS[3]], [minutes_listRTS[3]],[service_type_listRTS[3]]],
        [[dep_time_listRTS[4]], [minutes_listRTS[4]],[service_type_listRTS[4]]],
        [[dep_time_listRTS[5]], [minutes_listRTS[5]],[service_type_listRTS[5]]],
        [[dep_time_listRTS[6]], [minutes_listRTS[6]],[service_type_listRTS[6]]]

        ]
        
    #print(data)

    # adding the data extracted to the table using loop function
    global count
    count=0
    for record in data:
    
        set.insert(parent='',index='end',iid = count,text='',values=(record[0],record[1],record[2]))
    
        count += 1 #take the next line when added


    #only for DE
    rangeDE = ws._cells_by_col(8,1,10,10000)
    for row in rangeDE:
        for cell in row:
            if (cell.value == search_str.upper()):
                # FOR RTS2 DE
                def RTS2():
                    #setting up the table to display the result
                    ttk.Treeview(frame2)
                    set1 = ttk.Treeview(frame2)
                    set1.grid(row=7, column=0, columnspan=9, rowspan=2, pady=10, padx=10) #where the table will be located

                    set1['columns']= ('dispatch_time1', 'minutes1','service_type1') #columns IDs
                    set1.column("#0", width=0,  stretch=NO)
                    set1.column("dispatch_time1",anchor=CENTER, width=140)
                    set1.column("minutes1",anchor=CENTER, width=80)
                    set1.column("service_type1",anchor=CENTER, width=180)
                    

                    #name and formatting for headers
                    set1.heading("#0",text="",anchor=CENTER)
                    set1.heading("dispatch_time1",text="Dispatch Time",anchor=CENTER)
                    set1.heading("minutes1",text="Minutes",anchor=CENTER)
                    set1.heading("service_type1",text="Service Type",anchor=CENTER)
                    

                    #data in the table
                    data1  = [
                        [[dep_time_listRTS2[0]], [minutes_listRTS2[0]],[service_type_listRTS2[0]]],
                        [[dep_time_listRTS2[1]], [minutes_listRTS2[1]],[service_type_listRTS2[1]]],
                        [[dep_time_listRTS2[2]], [minutes_listRTS2[2]],[service_type_listRTS2[2]]],
                        [[dep_time_listRTS2[3]], [minutes_listRTS2[3]],[service_type_listRTS2[3]]],
                        [[dep_time_listRTS2[4]], [minutes_listRTS2[4]],[service_type_listRTS2[4]]],
                        [[dep_time_listRTS2[5]], [minutes_listRTS2[5]],[service_type_listRTS2[5]]],
                        [[dep_time_listRTS2[6]], [minutes_listRTS2[6]],[service_type_listRTS2[6]]]
                        
                        ]
                        
                    print(data1)

                
                #adding the data extracted to the table using loop function
                    global count1
                    count1=0
                    for record1 in data1:
                
                        set1.insert(parent='',index='end',iid = count1,text='',values=(record1[0],record1[1],record1[2]))
                
                        count1 += 1 #take the next line when added
                def RTS1():
                    set = ttk.Treeview(frame2)
                    set.grid(row=7, column=0, columnspan=9, rowspan=2, pady=10, padx=10) #where the table will be located

                    set['columns']= ('dispatch_time', 'minutes','service_type') #columns IDs
                    set.column("#0", width=0,  stretch=NO)
                    set.column("dispatch_time",anchor=CENTER, width=140)
                    set.column("minutes",anchor=CENTER, width=80)
                    set.column("service_type",anchor=CENTER, width=180)
                    

                    #name and formatting for headers
                    set.heading("#0",text="",anchor=CENTER)
                    set.heading("dispatch_time",text="Dispatch Time",anchor=CENTER)
                    set.heading("minutes",text="Minutes",anchor=CENTER)
                    set.heading("service_type",text="Service Type",anchor=CENTER)
                    

                    #data in the table
                    data  = [
                        [[dep_time_listRTS[0]], [minutes_listRTS[0]],[service_type_listRTS[0]]],
                        [[dep_time_listRTS[1]], [minutes_listRTS[1]],[service_type_listRTS[1]]],
                        [[dep_time_listRTS[2]], [minutes_listRTS[2]],[service_type_listRTS[2]]],
                        [[dep_time_listRTS[3]], [minutes_listRTS[3]],[service_type_listRTS[3]]],
                        [[dep_time_listRTS[4]], [minutes_listRTS[4]],[service_type_listRTS[4]]],
                        [[dep_time_listRTS[5]], [minutes_listRTS[5]],[service_type_listRTS[5]]],
                        [[dep_time_listRTS[6]], [minutes_listRTS[6]],[service_type_listRTS[6]]]

                        ]
                    global count
                    count=0
                    for record in data:
                    
                        set.insert(parent='',index='end',iid = count,text='',values=(record[0],record[1],record[2]))
                    
                        count += 1 #take the next line when added
                    

                RTS2button = tk.Button(frame2, text="  RTS2   ", padx=30, pady=3, fg="black", bg= "light grey", command= RTS2)
                RTS2button.grid(row=12, column=5, columnspan=2)
                
                RTS1button = tk.Button(frame2, text="  RTS1   ", padx=30, pady=3, fg="black", bg= "light grey", command= RTS1)
                RTS1button.grid(row=12, column=2, columnspan=2)
    

#For prescheduled RTS 1 blocks UK and DE
    rangePrescheduledRTS1 = ws._cells_by_col(14,1,16,10000)
    for row in rangePrescheduledRTS1:
        for cell in row:
            if (cell.value == search_str.upper()):
                # FOR RTS2 DE
                def PrescheduledRTS1():
                    #setting up the table to display the result
                    ttk.Treeview(frame2)
                    set1 = ttk.Treeview(frame2)
                    set1.grid(row=7, column=0, columnspan=9, rowspan=2, pady=10, padx=10) #where the table will be located

                    set1['columns']= ('dispatch_time1', 'minutes1','service_type1') #columns IDs
                    set1.column("#0", width=0,  stretch=NO)
                    set1.column("dispatch_time1",anchor=CENTER, width=140)
                    set1.column("minutes1",anchor=CENTER, width=80)
                    set1.column("service_type1",anchor=CENTER, width=180)
                    

                    #name and formatting for headers
                    set1.heading("#0",text="",anchor=CENTER)
                    set1.heading("dispatch_time1",text="Block Length",anchor=CENTER)
                    set1.heading("minutes1",text="Minutes",anchor=CENTER)
                    set1.heading("service_type1",text="Service Type",anchor=CENTER)
                    

                    #data in the table
                    data1  = [
                        [[dep_time_listRTS1prescheduled[0]], [minutes_listRTS1prescheduled[0]],[service_type_listRTS2[0]]],
                        [[dep_time_listRTS1prescheduled[1]], [minutes_listRTS1prescheduled[1]],[service_type_listRTS2[1]]],
                        [[dep_time_listRTS1prescheduled[2]], [minutes_listRTS1prescheduled[2]],[service_type_listRTS2[2]]],
                        [[dep_time_listRTS1prescheduled[3]], [minutes_listRTS1prescheduled[3]],[service_type_listRTS2[3]]],
                        [[dep_time_listRTS1prescheduled[4]], [minutes_listRTS1prescheduled[4]],[service_type_listRTS2[4]]],
                        [[dep_time_listRTS1prescheduled[5]], [minutes_listRTS1prescheduled[5]],[service_type_listRTS2[5]]],
                        [[dep_time_listRTS1prescheduled[6]], [minutes_listRTS1prescheduled[6]],[service_type_listRTS2[6]]]
                        
                        ]
                        
                    print(data1)

                
                #adding the data extracted to the table using loop function
                    global count1
                    count1=0
                    for record1 in data1:
                
                        set1.insert(parent='',index='end',iid = count1,text='',values=(record1[0],record1[1],record1[2]))
                
                        count1 += 1 #take the next line when added
                
                Spaceprescheduled = Label(frame2, text = "       ").grid(row=14,column=1, columnspan=7)
                PrescheduledRTS1button = tk.Button(frame2, text="Presched. RTS1", padx=40, pady=3, fg="red", bg= "white", command= PrescheduledRTS1)
                PrescheduledRTS1button.grid(row=15, column=2, columnspan=2)        
    
    rangePrescheduledRTS2 = ws._cells_by_col(19,1,21,10000)
    for row in rangePrescheduledRTS2:
        for cell in row:
            if (cell.value == search_str.upper()):                           
                def PrescheduledRTS2():
                    set = ttk.Treeview(frame2)
                    set.grid(row=7, column=0, columnspan=9, rowspan=2, pady=10, padx=10) #where the table will be located

                    set['columns']= ('dispatch_time', 'minutes','service_type') #columns IDs
                    set.column("#0", width=0,  stretch=NO)
                    set.column("dispatch_time",anchor=CENTER, width=140)
                    set.column("minutes",anchor=CENTER, width=80)
                    set.column("service_type",anchor=CENTER, width=180)
                    

                    #name and formatting for headers
                    set.heading("#0",text="",anchor=CENTER)
                    set.heading("dispatch_time",text="Block Length",anchor=CENTER)
                    set.heading("minutes",text="Minutes",anchor=CENTER)
                    set.heading("service_type",text="Service Type",anchor=CENTER)
                    

                    #data in the table
                    data  = [
                        [[dep_time_listRTS2prescheduled[0]], [minutes_listRTS2prescheduled[0]],[service_type_listRTS[0]]],
                        [[dep_time_listRTS2prescheduled[1]], [minutes_listRTS2prescheduled[1]],[service_type_listRTS[1]]],
                        [[dep_time_listRTS2prescheduled[2]], [minutes_listRTS2prescheduled[2]],[service_type_listRTS[2]]],
                        [[dep_time_listRTS2prescheduled[3]], [minutes_listRTS2prescheduled[3]],[service_type_listRTS[3]]],
                        [[dep_time_listRTS2prescheduled[4]], [minutes_listRTS2prescheduled[4]],[service_type_listRTS[4]]],
                        [[dep_time_listRTS2prescheduled[5]], [minutes_listRTS2prescheduled[5]],[service_type_listRTS[5]]],
                        [[dep_time_listRTS2prescheduled[6]], [minutes_listRTS2prescheduled[6]],[service_type_listRTS[6]]]

                        ]
                    global count
                    count=0
                    for record in data:
                    
                        set.insert(parent='',index='end',iid = count,text='',values=(record[0],record[1],record[2]))
                    
                        count += 1 #take the next line when added
                    
                Spaceprescheduled = Label(frame2, text = "       ").grid(row=14,column=1, columnspan=7)
                PrescheduledRTS2button = tk.Button(frame2, text="Presched. RTS2", padx=40, pady=3, fg="red", bg= "white", command= PrescheduledRTS2)
                PrescheduledRTS2button.grid(row=15, column=5, columnspan=2)
                


    #looking for flex SIMs (if Any)
    wRTS = wb.worksheets[2]
    rangeexcel = wRTS.iter_rows() # Defaults to whole sheet (to search)
    for row in rangeexcel:
        for cell in row:
            if (cell.value == search_str.upper()):
                #take the necessary info from excel
                global SIMRTS
                global SIMRTSinfo
                SIMRTS = wRTS.cell(row=cell.row, column=3).value
                SIMRTSinfo = wRTS.cell(row=1, column=3).value
                #Sentence informing there is a SIM
                
                                      
                SIMRTSinfolabel = Label(frame0,  text= SIMRTSinfo, justify=LEFT, font = "Helvetica 10", foreground = "red")
                SIMRTSinfolabel.grid(row=13, column=0, columnspan=1, pady=10, padx=10)
                def callback(url):
                    webbrowser.open_new_tab(url)
                SIMRTSlabel = Label(frame0,  text= SIMRTS, justify=LEFT, cursor="hand2", font = "Helvetica 10 underline", foreground = "blue")
                SIMRTSlabel.grid(row=15, column=3, columnspan=1, pady=10, padx=10)
                SIMRTSlabel.bind("<Button-1>", lambda e: callback(SIMRTS))
            

# function for when clicking Adhoc button
#it is a copy paste from the RTS function with slight modifications to meet the criteria, therefore no explanations needed
def excelfileAdhoc():
    top1 = Toplevel()
    top1.title()
    global minutesADHOC #it must be a global variable to be displayed otherwise it wont work
    global minutes_listADHOC
    global dep_time_listADHOC

    global DS
    DS = DSentry.get()
    DSentry.delete(0, END)
    DSupper = DS.upper()
    top1.title(DSupper)

    
    #title in the new window ( RTS + DS)
    frametitleAdhoc = LabelFrame(top1, padx=1, pady=1)
    frametitleAdhoc.grid(row=4, columnspan=11, \
                 padx=1, pady=1, ipadx=1, ipady=1)
    titlelabelAdhoc = Label(frametitleAdhoc, text = "Adhoc " + DSupper, anchor= CENTER,  font = "Helvetica 15", padx=1, pady=1).grid(row=4, column=0, columnspan=1, pady=1, padx=1)

    #Setting frames for DSP minutes and Flex minutes
    #for flex
    frame2adhoc = LabelFrame(top1, padx=5, pady=5, text="Flex", font = "16")
    frame2adhoc.grid(row=5, columnspan=11, \
                 padx=5, pady=5, ipadx=5, ipady=5)
    #for DSP
    frame1adhoc = LabelFrame(top1, padx=5, pady=5, text="DSP", font = "16") 
    frame1adhoc.grid(row=11, columnspan=11, \
                 padx=5, pady=5, ipadx=5, ipady=5)

    #for SIMs
    frame0adhoc = LabelFrame(top1, padx=5, pady=5, text="SIMs", font = "16") 
    frame0adhoc.grid(row=13, columnspan=11, \
                 padx=5, pady=5, ipadx=5, ipady=5)


    #Reading Excel file
    
    # setting directory where file is located
    path = "//ant/dept-eu/TBA/UK/Business Analyses/CentralOPS/PM Shift/DHP1/PM_Dashboard/Sequencing inputs/Database.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[1] #sheet number 1 is "RTS"
    wsES = wb.worksheets[4]

    #looking for minutes depending on dep time            
    search_str = DS #we want to search for the DS specified
    minutes_listADHOC = []
    dep_time_listADHOC = []
    service_type_listADHOC = [] #the output of the minutes in list format
    range1 = ws.iter_rows() # Defaults to whole sheet (to search)
    range2 = wsES.iter_rows()
    
    
    for row in range2:
        for cell in row:
            if (cell.value == search_str.upper()):
                minutesAdhoc0 = ws.cell(row=6, column=11).value #get the value from column 7 and the row where DS has been found
                minutes_listADHOC.append(minutesAdhoc0)
                           
                deptimesAdhoc0 = ws.cell(row=6, column=10).value
                dep_time_listADHOC.append(deptimesAdhoc0)

                servicetypeAdhoc0 = ws.cell(row=6, column=12).value
                service_type_listADHOC.append(servicetypeAdhoc0)
            
            else:
                for row in range1:
                    for cell in row:
                        if (cell.value == search_str.upper()):
                            minutesADHOC = ws.cell(row=cell.row, column=7).value
                            minutes_listADHOC.append(minutesADHOC)
                            #print(minutes_listADHOC)
                            
                            deptimesADHOC = ws.cell(row=cell.row, column=6).value
                            dep_time_listADHOC.append(deptimesADHOC)
                            #print(dep_time_listADHOC)

                            servicetypeADHOC = ws.cell(row=cell.row, column=4).value
                            service_type_listADHOC.append(servicetypeADHOC)
                            #print(service_type_list)
                



    set = ttk.Treeview(frame2adhoc)
    set.grid(row=6, column=0, columnspan=9, rowspan=2, pady=10, padx=10)

    set['columns']= ('dispatch_time', 'minutes','service_type') #columns IDs
    set.column("#0", width=0,  stretch=NO)
    set.column("dispatch_time",anchor=CENTER, width=150)
    set.column("minutes",anchor=CENTER, width=80)
    set.column("service_type",anchor=CENTER, width=140)
    

    #name and formatting for headers
    set.heading("#0",text="",anchor=CENTER)
    set.heading("dispatch_time",text="Dispatch Time",anchor=CENTER)
    set.heading("minutes",text="Minutes",anchor=CENTER)
    set.heading("service_type",text="Service Type",anchor=CENTER)

    #data to be taken. In this case we only have one value for every one of the lists generated (minutes, service type, dep time)
    data  = [
        [[dep_time_listADHOC[0]], [minutes_listADHOC[0]],[service_type_listADHOC[0]]]
        
        ]
        
    print(data)

    global count
    count=0
    for record in data:
    
        set.insert(parent='',index='end',iid = count,text='',values=(record[0],record[1],record[2]))
    
        count += 1   

    
    
    


    #looking for SIMs (if Any)
    wAdhoc = wb.worksheets[3]
    rangeexcel = wAdhoc.iter_rows() # Defaults to whole sheet (to search)
    for row in rangeexcel:
        for cell in row:
            if (cell.value == search_str.upper()):
                #take the necessary info from excel
                SIMlink = wAdhoc.cell(row=cell.row, column=3).value
                SIMAdhocinfo = wAdhoc.cell(row=1, column=3).value
                #Sentence informing there is a SIM
                SIMAdhocinfolabel = Label(frame0adhoc,  text= SIMAdhocinfo, justify=LEFT, font = "Helvetica 10", foreground = "red")
                SIMAdhocinfolabel.grid(row=13, column=0, columnspan=1, pady=10, padx=10)
                            
                def openurl(url):
                    webbrowser.open_new_tab(url)
                            
                SIMAdhoclabel = Label(frame0adhoc,  text= SIMlink, justify=LEFT, cursor="hand2", font = "Helvetica 10 underline", foreground = "blue")
                SIMAdhoclabel.grid(row=13, column=3, columnspan=1, pady=10, padx=10)
                SIMAdhoclabel.bind("<Button-1>", lambda e: openurl(SIMlink))
            
                
    #looking for DSP information
    
    AdhocDSP = ws.cell(row=2, column=4).value
    def openurl(url):
                    webbrowser.open_new_tab(url)
    AdhocDSPlabel = Label(frame1adhoc,  text= "ACES Standard", justify=LEFT, cursor="hand2", font = "Helvetica 10 underline", foreground = "blue")
    AdhocDSPlabel.grid(row=11, column=2, columnspan=1, pady=10, padx=10)
    AdhocDSPlabel.bind("<Button-1>", lambda e: openurl(AdhocDSP))

    SIMAdhocDSPinfo = ws.cell(row=1, column=4).value
    SIMAdhocinfolabel = Label(frame1adhoc,  text= SIMAdhocDSPinfo, justify=LEFT, font = "Helvetica 10")
    SIMAdhocinfolabel.grid(row=11, column=0, columnspan=1, pady=10, padx=10)

def buffers():
    top1 = Toplevel()
    top1.title()
    global buffers 
    global DS
    DS = DSentry.get()
    DSentry.delete(0, END)
    DSupper = DS.upper()
    top1.title(DSupper)

    
    #title in the new window ( RTS + DS)
    frametitlebuffer = LabelFrame(top1, padx=1, pady=1)
    frametitlebuffer.grid(row=0, columnspan=11, \
                 padx=1, pady=1, ipadx=1, ipady=1)
    titlelabelbuffer = Label(frametitlebuffer, text = "Buffers " + DSupper, anchor= CENTER,  font = "Helvetica 15", padx=1, pady=1).grid(row=4, column=0, columnspan=1, pady=1, padx=1)

    #Setting frames for DSP minutes and Flex minutes
    #for flex
    framebuffer = LabelFrame(top1, padx=5, pady=5)
    framebuffer.grid(row=5, columnspan=11, \
                 padx=5, pady=5, ipadx=5, ipady=5)
    
    
    
    
    #Extracting buffer
    # setting directory where file is located
    path = "//ant/dept-eu/TBA/UK/Business Analyses/CentralOPS/PM Shift/DHP1/PM_Dashboard/Sequencing inputs/Database.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[5] #sheet number 1 is "RTS"
    

    #looking for minutes depending on dep time            
    search_str = DS #we want to search for the DS specified
    
    rangeSSDA = ws._cells_by_col(2,1,2,10000) # Defaults to whole sheet (to search)
    for row in rangeSSDA:
        for cell in row:
            if (cell.value == search_str.upper()):
                SSDAbuffer = ws.cell(row=cell.row, column=4).value
                print(SSDAbuffer)
    try: SSDAbuffer
    except NameError: SSDAbuffer = None
    if SSDAbuffer is None:
        SSDAbuffer = 0


    rangeSSDB = ws._cells_by_col(7,1,7,10000)
    for row in rangeSSDB:
        for cell in row:
            if (cell.value == search_str.upper()):
                SSDBbuffer = ws.cell(row=cell.row, column=9).value
                print(SSDBbuffer)
            
    try: SSDBbuffer
    except NameError: SSDBbuffer = None
    if SSDBbuffer is None:
        SSDBbuffer = 0


    rangeSSDC = ws._cells_by_col(12,1,12,1000)
    for row in rangeSSDC:
        for cell in row:
            if (cell.value == search_str.upper()):
                SSDCbuffer = ws.cell(row=cell.row, column=14).value
                print(SSDCbuffer)
                
    try: SSDCbuffer
    except NameError: SSDCbuffer = None
    if SSDCbuffer is None:
        SSDCbuffer = 0
    
    
    
    
    #Calculating the total
    def calculate():
        global SSDAtotal
        global SSDBtotal
        global SSDCtotal
        global SSDA
        global SSDB
        global SSDC

        SSDA = SSDAentry.get()
        SSDB = SSDBentry.get()
        SSDC = SSDCentry.get()

        if SSDA == "":
            SSDA = 0
        if SSDB == "":
            SSDB = 0
        if SSDC == "":
            SSDC = 0

        SSDAint = float(SSDA)
        SSDBint = float(SSDB)
        SSDCint = float(SSDC)



        SSDAbufferint = float(SSDAbuffer)
        SSDBbufferint = float(SSDBbuffer)
        SSDCbufferint = float(SSDCbuffer)



        SSDAtotal = round(decimal.Decimal(SSDAint*SSDAbufferint),2)
        SSDBtotal =  round(decimal.Decimal(SSDBint*SSDBbufferint),2)
        SSDCtotal = round(decimal.Decimal(SSDCint*SSDCbufferint),2)

        

       

        Total = Label(framebuffer, text = "Total",font = "Helvetica 12").grid(row=1,column=6, columnspan=1)
        TotalSSDA = Label(framebuffer, text = SSDAtotal).grid(row=2,column=6, columnspan=1)
        TotalSSDB = Label(framebuffer, text = SSDBtotal).grid(row=3,column=6, columnspan=1)
        TotalSSDC = Label(framebuffer, text = SSDCtotal).grid(row=4,column=6, columnspan=1)
        #Space"c"
        Spacec = Label(framebuffer, text = "       ",font = "Helvetica 12").grid(row=1,column=5, columnspan=1)
        Space1c = Label(framebuffer, text = "       ").grid(row=2,column=5, columnspan=1)
        Space2c = Label(framebuffer, text = "       ").grid(row=3,column=5, columnspan=1)
        Space3c = Label(framebuffer, text = "       ").grid(row=4,column=5, columnspan=1)
    
    
    Cycle = Label(framebuffer, text = "Cycle", font = "Helvetica 12").grid(row=1,column=0, columnspan=1)
    CycleSSDA = Label(framebuffer, text = "SSDA").grid(row=2,column=0, columnspan=1)
    CycleSSDB = Label(framebuffer, text = "SSDB").grid(row=3,column=0, columnspan=1)
    CycleSSDC = Label(framebuffer, text = "SSDC").grid(row=4,column=0, columnspan=1)
    
    
    SSDAbufferint = round((float(SSDAbuffer)*100),2)
    SSDBbufferint = round((float(SSDBbuffer)*100),2)
    SSDCbufferint = round((float(SSDCbuffer)*100),2)
    
    Buffer = Label(framebuffer, text = "Buffer",font = "Helvetica 12").grid(row=1,column=2, columnspan=1)
    BufferSSDA = Label(framebuffer, text = str(SSDAbufferint) +" %").grid(row=2,column=2, columnspan=1)
    BufferSSDB = Label(framebuffer, text = str(SSDBbufferint) +" %").grid(row=3,column=2, columnspan=1)
    BufferSSDC = Label(framebuffer, text = str(SSDCbufferint) +" %").grid(row=4,column=2, columnspan=1)

    
    #Ask how many blocks generated
    Amount = Label(framebuffer, text = "Blocks generated",font = "Helvetica 12").grid(row=1,column=4, columnspan=1)
    SSDAentry = Entry(framebuffer, width=10, borderwidth=5)
    SSDAentry.grid(row=2, column=4, columnspan=1, pady=10, padx=10)
    SSDBentry = Entry(framebuffer, width=10, borderwidth=5)
    SSDBentry.grid(row=3, column=4, columnspan=1, pady=10, padx=10)
    SSDCentry = Entry(framebuffer, width=10, borderwidth=5)
    SSDCentry.grid(row=4, column=4, columnspan=1, pady=10, padx=10)

    if SSDAbufferint == 0:
        SSDAentry.config(state="disabled")
    if SSDBbufferint == 0:
        SSDBentry.config(state="disabled")
    if SSDCbufferint == 0:
        SSDCentry.config(state="disabled")

    #Space"a"
    Spacea = Label(framebuffer, text = "       ",font = "Helvetica 12").grid(row=1,column=1, columnspan=1)
    Space1a = Label(framebuffer, text = "       ").grid(row=2,column=1, columnspan=1)
    Space2a = Label(framebuffer, text = "       ").grid(row=3,column=1, columnspan=1)
    Space3a = Label(framebuffer, text = "       ").grid(row=4,column=1, columnspan=1)
    #Space"b"
    Spaceb = Label(framebuffer, text = "       ",font = "Helvetica 12").grid(row=1,column=3, columnspan=1)
    Space1b = Label(framebuffer, text = "       ").grid(row=2,column=3, columnspan=1)
    Space2b = Label(framebuffer, text = "       ").grid(row=3,column=3, columnspan=1)
    Space3b = Label(framebuffer, text = "       ").grid(row=4,column=3, columnspan=1)
    

    Bufferscalculationbutton = tk.Button(top1, text="  Calculate   ", padx=30, pady=3, fg="black", bg= "light grey", command= calculate)
    Bufferscalculationbutton.grid(row=8, column=4, columnspan=2)
    Spacebuttonabove = Label(top1, text = "       ", padx=30, pady=1).grid(row=7,column=4, columnspan=1)
    Spacebuttonbelow = Label(top1, text = "       ", padx=30, pady=1).grid(row=9,column=4, columnspan=1)

    #Note for the scheduler
    Note = ws.cell(row=9, column=17).value
    Notebuffer = Label(top1, text = Note, padx=30, pady=1).grid(row=10,column=4, columnspan=1)
    Notebufferspace = Label(top1, text = "", padx=30, pady=1).grid(row=11,column=4, columnspan=1)

#
def addDS(DSnodecountry):
    DSentry.delete(0,END)
    DSentry.insert(0,DSnodecountry)

#calling the module checklist from Michal
def checklist():
    checklistmain.main_script_checklist(DSentry.get().upper())
    DSentry.delete(0,END)

#calling the module sequencing report from Michal
def sequencingreport():
    def submitinfo():
        OFDwindow.destroy()
        SequencingReport.main_script_sequencingreport(DSentry.get().upper(),variable.get(),True)
        DSentry.delete(0,END)
        

        
        
    OFDwindow = Toplevel(root)
    tomorrowOFD = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
    sequencinginputsOFD2= (datetime.now() + timedelta(days=2)).strftime("%Y-%m-%d")
    sequencinginputsOFD3= (datetime.now() + timedelta(days=3)).strftime("%Y-%m-%d")
    sequencinginputsOFD4= (datetime.now() + timedelta(days=4)).strftime("%Y-%m-%d")
    Options = [tomorrowOFD, sequencinginputsOFD2, sequencinginputsOFD3, sequencinginputsOFD4]
    variable = tk.StringVar(root)
    variable.set(Options[0])
    OFDdate = tk.OptionMenu(OFDwindow, variable,*Options)
    OFDdate.config(width=20, borderwidth=1)
    OFDdate.grid(row=2, column=1, columnspan=1, pady=10, padx=5)

    DSselectedlabel = Label(OFDwindow,  text= DSentry.get().upper(), justify=CENTER, font = "Helvetica 15", padx=20, pady=5)
    DSselectedlabel.grid(row=1, column=0, columnspan=2, pady=10, padx=10)
    textlabelOFD = Label(OFDwindow,  text= "OFD date:", justify=LEFT, padx=20, pady=5)
    textlabelOFD.grid(row=2, column=0, columnspan=1, pady=10, padx=10)
    buttonOFDsubmit = tk.Button(OFDwindow, text = "Select", padx=44, pady=3, fg="black", bg= "light grey", command= submitinfo)
    buttonOFDsubmit.grid(row=3, column=0, columnspan=2, pady=10, padx=10)


#define labels and DS input
spacetextabove = Label(root, text = "").grid(row=0,column=0, columnspan=4)
spacetextbelow = Label(root, text = "").grid(row=4,column=0, columnspan=4)
spacetextbelow1 = Label(root, text = "", padx=20, pady=5 ).grid(row=3,column=0, columnspan=1)
textlabelDS = Label(root,  text= "Delivery Station:", justify=LEFT, padx=20, pady=5).grid(row=1, column=0, columnspan=1, pady=10, padx=10)
DSentry = Entry(root, width=45, borderwidth=5)
DSentry.grid(row=1, column=1, columnspan=3, pady=10, padx=10)

#definiing the buttons alongside attributing the functions/commands
RTSbutton = tk.Button(root, text=" RTS  ", padx=44, pady=3, fg="black", bg= "light grey", command= excelfileRTS)
RTSbutton.grid(row=3, column=0, columnspan=1)

Adhocbutton = tk.Button(root, text="  Adhoc   ", padx=35, pady=3, fg="black", bg= "light grey", command= excelfileAdhoc)
Adhocbutton.grid(row=3, column=1, columnspan=1)

Buffersbutton = tk.Button(root, text="  SD Buffers   ", padx=30, pady=3, fg="black", bg= "light grey", command= buffers)
Buffersbutton.grid(row=3, column=2, columnspan=1)

Checklistbutton = tk.Button(root, text=" Checklist  ", padx=30, pady=3, fg="black", bg= "light grey", command= checklist)
Checklistbutton.grid(row=20, column=0, columnspan=2)

SequencingReportbutton = tk.Button(root, text=" Sequencing Report  ", padx=10, pady=3, fg="black", bg= "light grey", command= sequencingreport)
SequencingReportbutton.grid(row=20, column=1, columnspan=2)

Checklistbuttonspace = Label(root, text=" ")
Checklistbuttonspace.grid(row=19, column=0, columnspan=3)

Checklistbuttonspace1 = Label(root, text=" ")
Checklistbuttonspace1.grid(row=21, column=0, columnspan=3)




#finding scheduler's DS and adding to root
try:
    dateVar = datetime.today()
    dateStr1 = dateVar.strftime('%d-%m-%Y')
    dateStr2 = dateVar.strftime('%Y-%m-%d')
    tasklistlocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 + '\\TaskList ' + dateStr1 + '.xlsx'
    wb = openpyxl.load_workbook(tasklistlocation) #load workbook and telling Python to read only
    search_str = user_login
    wb = openpyxl.load_workbook(tasklistlocation) #load workbook and telling Python to read only
    ws = wb.worksheets[0] #sheet number 1
    rangeexcel = ws.iter_rows()

    DSschedulerlist1 = []
    DSschedulerlist2 = []
    DSschedulerlist3 = []
    for row in rangeexcel:
            for cell in row:
                if (cell.value == search_str):
                    DSnode1buddy1 = ws.cell(row=cell.row, column=4).value #get the value from column 4 and the row where DS has been found
                    DSnode2buddy1 = ws.cell(row=cell.row, column=5).value
                    DSnode3buddy1 = ws.cell(row=cell.row, column=6).value
                    DSnode4buddy1 = ws.cell(row=cell.row, column=7).value
                    DSnode5buddy1 = ws.cell(row=cell.row, column=8).value
                    DSnode6buddy1 = ws.cell(row=cell.row, column=9).value
                    DSnode7buddy1 = ws.cell(row=cell.row, column=10).value
                    DSnode8buddy1 = ws.cell(row=cell.row, column=11).value
                    DSnode9buddy1 = ws.cell(row=cell.row, column=12).value
                    DSnode10buddy1 = ws.cell(row=cell.row, column=13).value
                    DSnode11buddy1 = ws.cell(row=cell.row, column=14).value
                    DSnode12buddy1 = ws.cell(row=cell.row, column=15).value
                    DSnode13buddy1 = ws.cell(row=cell.row, column=16).value
                    DSnode14buddy1 = ws.cell(row=cell.row, column=17).value
                    DSnode15buddy1 = ws.cell(row=cell.row, column=18).value
                    DSnode16buddy1 = ws.cell(row=cell.row, column=19).value
                    DSnode17buddy1 = ws.cell(row=cell.row, column=20).value
                    DSnode18buddy1 = ws.cell(row=cell.row, column=21).value
                    DSnode19buddy1 = ws.cell(row=cell.row, column=22).value
                    DSnode20buddy1 = ws.cell(row=cell.row, column=23).value
                    DSnode21buddy1 = ws.cell(row=cell.row, column=24).value
                    DSnode22buddy1 = ws.cell(row=cell.row, column=25).value
                    DSnode23buddy1 = ws.cell(row=cell.row, column=26).value
                    DSnode24buddy1 = ws.cell(row=cell.row, column=27).value
                    
    #adding DS in 3 lists, 1 for each column to display them
    DSschedulerlist1.extend((DSnode1buddy1,DSnode4buddy1,DSnode7buddy1,DSnode10buddy1,DSnode13buddy1,DSnode16buddy1,DSnode19buddy1,DSnode22buddy1))
    DSschedulerlist2.extend((DSnode2buddy1,DSnode5buddy1,DSnode8buddy1,DSnode11buddy1,DSnode14buddy1,DSnode17buddy1,DSnode20buddy1,DSnode23buddy1))
    DSschedulerlist3.extend((DSnode3buddy1,DSnode6buddy1,DSnode9buddy1,DSnode12buddy1,DSnode15buddy1,DSnode18buddy1,DSnode21buddy1,DSnode24buddy1))

    shiftconfiglocation = r'\\ant\dept-eu\TBA\UK\Business Analyses\CentralOPS\PM Shift\DHP1\TaskList\TasklistGeneration\\' + dateStr2 + '\\shift_config.xlsx'
    wconfig = openpyxl.load_workbook(shiftconfiglocation)


    #because the DS are retrieved with SD shortcode next to the DS code e.g "DNX1 - SSD", we just want the DS code
    wconfigStations= wconfig.worksheets[1]
    rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)            
    DSscheduler1country1 = []
    for x in DSschedulerlist1:
        if x != None:
            DSscheduler1country1.append(x[0:4]) 

    rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)  
    rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)            
    DSscheduler1country2 = []
    for x in DSschedulerlist2:
        if x != None:
            DSscheduler1country2.append(x[0:4]) 

    rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)  
    rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)            
    DSscheduler1country3 = []
    for x in DSschedulerlist3:
        if x != None:
            DSscheduler1country3.append(x[0:4]) 

    #now we want to see from which country is every DS
    countrylist1 = []
    countrylist2 = []
    countrylist3 = []


    for DScountrycode in DSscheduler1country1:
        wconfigStations= wconfig.worksheets[1]
        rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
        try:    
            for row in rangeconfig:
                for cell in row:
                    if (cell.value == DScountrycode):
                        countrylist1.append(wconfigStations.cell(row=cell.row, column=2).value)
        except Exception: 
            pass


    for DScountrycode in DSscheduler1country2:
        wconfigStations= wconfig.worksheets[1]
        rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
        try:    
            for row in rangeconfig:
                for cell in row:
                    if (cell.value == DScountrycode):
                        countrylist2.append(wconfigStations.cell(row=cell.row, column=2).value)
        except Exception: 
            pass


    for DScountrycode in DSscheduler1country3:
        wconfigStations= wconfig.worksheets[1]
        rangeconfig = wconfigStations._cells_by_col(1,1,2,10000)
        try:    
            for row in rangeconfig:
                for cell in row:
                    if (cell.value == DScountrycode):
                        countrylist3.append(wconfigStations.cell(row=cell.row, column=2).value)
        except Exception: 
            pass


    #displaying the DS and giving them a colour in terms of country
    #colour formatting and displaying it
    global count
    count = 7 
    colourlist = []

    for x,y in zip(countrylist1,DSschedulerlist1):
        print(x)
        #print(y)
        if x == "UK":
            colour="orange"
        if x == "FR":
            colour="cyan"
        if x == "DE":
            colour="yellow"
        if x == "ES":
            colour="magenta"
        if x == "IT":
            colour="green"
        if x == "BE":
            colour = "SkyBlue4"
        if x == "NL":
            colour = "pale green"
        if x == "AT":
            colour = "light yellow"
        
        
        colourlist.append(colour)
        
    ButtonDS = {}
    colourlistsize = list(range(len(colourlist)))
    print(colourlistsize)
    for DSnumber,DSnodecountry,DSnode in zip(colourlistsize,DSscheduler1country1,DSschedulerlist1):
        action = lambda x = DSnodecountry: addDS(x)
        ButtonDS[DSnodecountry]= tk.Button(root, text = DSnode ,padx=20, pady=2, relief=GROOVE, justify=CENTER,width=10, background= colourlist[DSnumber], command=action)
        ButtonDS[DSnodecountry].grid(row=(count), column=0, columnspan=1, pady=5, padx=10)
        count += 1


    global count2
    count2 = 7 
    colourlist2 = []

    for x,y in zip(countrylist2,DSschedulerlist2):
        print(x)
        #print(y)
        if x == "UK":
            colour="orange"
        if x == "FR":
            colour="cyan"
        if x == "DE":
            colour="yellow"
        if x == "ES":
            colour="magenta"
        if x == "IT":
            colour="green"
        if x == "BE":
            colour = "SkyBlue4"
        if x == "NL":
            colour = "pale green"
        if x == "AT":
            colour = "light yellow"
        
        
        colourlist2.append(colour)
        
    ButtonDS = {}
    colourlist2size = list(range(len(colourlist2)))
    print(colourlist2size)
    for DSnumber,DSnodecountry,DSnode in zip(colourlist2size,DSscheduler1country2,DSschedulerlist2):
        action = lambda x = DSnodecountry: addDS(x)
        ButtonDS[DSnodecountry]= tk.Button(root, text = DSnode ,padx=20, pady=2, relief=GROOVE, justify=CENTER,width=10, background= colourlist2[DSnumber], command=action)
        ButtonDS[DSnodecountry].grid(row=(count2), column=1, columnspan=1, pady=5, padx=10)
        count2 += 1


    global count3
    count3 = 7 
    colourlist3 = []

    for x,y in zip(countrylist3,DSschedulerlist3):
        print(x)
        #print(y)
        if x == "UK":
            colour="orange"
        if x == "FR":
            colour="cyan"
        if x == "DE":
            colour="yellow"
        if x == "ES":
            colour="magenta"
        if x == "IT":
            colour="green"
        if x == "BE":
            colour = "SkyBlue4"
        if x == "NL":
            colour = "pale green"
        if x == "AT":
            colour = "light yellow"
        
        
        colourlist3.append(colour)
        
    ButtonDS = {}
    colourlist3size = list(range(len(colourlist3)))
    print(colourlist3size)
    for DSnumber,DSnodecountry,DSnode in zip(colourlist3size,DSscheduler1country3,DSschedulerlist3):
        action = lambda x = DSnodecountry: addDS(x)
        ButtonDS[DSnodecountry]= tk.Button(root, text = DSnode ,padx=20, pady=2, relief=GROOVE, justify=CENTER,width=10, background= colourlist3[DSnumber], command=action)
        ButtonDS[DSnodecountry].grid(row=(count3), column=2, columnspan=1, pady=5, padx=10)
        count3 += 1



except:
    print("Scheduler/Manager has no DS allocated")




    



root.mainloop()